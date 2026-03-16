"""File parsing and chunking pipeline for RAG ingestion."""

from __future__ import annotations

import csv
import io
import uuid
from pathlib import Path

import chardet


# ── File Parsers ──────────────────────────────────────

def parse_txt(content: bytes) -> str:
    encoding = chardet.detect(content).get("encoding", "utf-8") or "utf-8"
    return content.decode(encoding, errors="replace")


def parse_pdf(content: bytes) -> str:
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(content))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)
    return "\n\n".join(pages)


def parse_docx(content: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def parse_xlsx(content: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sections = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            sections.append(f"## Sheet: {sheet.title}\n" + "\n".join(rows))
    return "\n\n".join(sections)


def parse_csv_file(content: bytes) -> str:
    encoding = chardet.detect(content).get("encoding", "utf-8") or "utf-8"
    text = content.decode(encoding, errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = []
    for row in reader:
        if any(cell.strip() for cell in row):
            rows.append(" | ".join(row))
    return "\n".join(rows)


def describe_image(filename: str) -> str:
    """Return a placeholder description for images. 
    With an API key, this could call a vision model."""
    return f"[Image file: {filename}]"


PARSERS: dict[str, callable] = {
    ".txt": parse_txt,
    ".md": parse_txt,
    ".log": parse_txt,
    ".json": parse_txt,
    ".xml": parse_txt,
    ".html": parse_txt,
    ".htm": parse_txt,
    ".pdf": parse_pdf,
    ".docx": parse_docx,
    ".xlsx": parse_xlsx,
    ".xls": parse_xlsx,
    ".csv": parse_csv_file,
    ".tsv": parse_csv_file,
}

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}


def parse_file(filename: str, content: bytes) -> str:
    """Parse a file into plain text based on its extension."""
    ext = Path(filename).suffix.lower()

    if ext in IMAGE_EXTENSIONS:
        return describe_image(filename)

    parser = PARSERS.get(ext)
    if parser:
        return parser(content)

    # Fallback: try as text
    try:
        return parse_txt(content)
    except Exception:
        return f"[Unsupported file format: {ext}]"


# ── Chunking ──────────────────────────────────────────

def chunk_text(
    text: str,
    chunk_size: int = 800,
    chunk_overlap: int = 200,
    separator: str = "\n\n",
) -> list[str]:
    """Split text into overlapping chunks using semantic boundaries."""
    if not text.strip():
        return []

    # First split by the separator (paragraph boundaries)
    sections = text.split(separator)
    sections = [s.strip() for s in sections if s.strip()]

    chunks: list[str] = []
    current_chunk = ""

    for section in sections:
        # If a single section is already too large, split it further
        if len(section) > chunk_size:
            if current_chunk:
                chunks.append(current_chunk.strip())
                current_chunk = ""

            words = section.split()
            sub_chunk = ""
            for word in words:
                if len(sub_chunk) + len(word) + 1 > chunk_size:
                    if sub_chunk:
                        chunks.append(sub_chunk.strip())
                        # Keep overlap
                        overlap_words = sub_chunk.split()[-chunk_overlap // 5 :]
                        sub_chunk = " ".join(overlap_words) + " "
                sub_chunk += word + " "
            if sub_chunk.strip():
                chunks.append(sub_chunk.strip())
            continue

        # Check if adding this section exceeds chunk_size
        candidate = (current_chunk + separator + section).strip() if current_chunk else section
        if len(candidate) > chunk_size and current_chunk:
            chunks.append(current_chunk.strip())
            # Overlap: keep the tail of the previous chunk
            overlap_text = current_chunk[-chunk_overlap:] if chunk_overlap > 0 else ""
            current_chunk = overlap_text + separator + section if overlap_text else section
        else:
            current_chunk = candidate

    if current_chunk.strip():
        chunks.append(current_chunk.strip())

    return chunks


# ── Ingest Pipeline ───────────────────────────────────

def ingest_file(
    filename: str,
    content: bytes,
    agent_id: str | None = None,
    department: str | None = None,
    chunk_size: int = 800,
    chunk_overlap: int = 200,
) -> dict:
    """Parse, chunk, and store a file into the appropriate vector DB collection.
    
    Priority: agent_id > department > shared company knowledge.
    Returns stats about the ingestion.
    """
    from backend.engine.memory import AgentMemory, DepartmentKnowledgeBase, SharedKnowledgeBase

    text = parse_file(filename, content)
    if not text or text.startswith("[Unsupported") or text.startswith("[Image"):
        chunks = [text] if text else []
    else:
        chunks = chunk_text(text, chunk_size=chunk_size, chunk_overlap=chunk_overlap)

    if not chunks:
        return {"filename": filename, "chunks": 0, "target": "none"}

    file_metadata = {
        "source": "file_upload",
        "filename": filename,
        "file_type": Path(filename).suffix.lower(),
    }

    if agent_id:
        memory = AgentMemory(agent_id)
        for i, chunk in enumerate(chunks):
            memory.store_semantic(
                chunk,
                metadata={**file_metadata, "chunk_index": str(i), "total_chunks": str(len(chunks))},
            )
        target = f"agent:{agent_id}"
    elif department:
        dept_kb = DepartmentKnowledgeBase(department)
        for i, chunk in enumerate(chunks):
            dept_kb.add(
                chunk,
                metadata={**file_metadata, "chunk_index": str(i), "total_chunks": str(len(chunks))},
            )
        target = f"department:{department}"
    else:
        kb = SharedKnowledgeBase()
        for i, chunk in enumerate(chunks):
            kb.add(
                chunk,
                metadata={**file_metadata, "chunk_index": str(i), "total_chunks": str(len(chunks))},
            )
        target = "shared_knowledge"

    return {
        "filename": filename,
        "chunks": len(chunks),
        "characters": len(text),
        "target": target,
    }
